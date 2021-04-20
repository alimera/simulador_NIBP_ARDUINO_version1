import openpyxl
import matplotlib.pyplot as plt
import matplotlib
import tkinter
matplotlib.use('TkAgg')
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
from tkinter import *
import serial, time

arduino = serial.Serial("COM4", 115200)
time.sleep(1)

##lectura de archivo
archivo ='C:/Users/aliva/Desktop/noveno/bio2\proyecto final/codigointento/NIBPDataSet.xlsx'
wb_obj = openpyxl.load_workbook(archivo)
sheet_obj = wb_obj.active
#m_row = sheet_obj.max_row
m_row=823
m_colum = sheet_obj.max_column


#Variables para graficar
PM=[]
tiempo=[]
PMr=[]


#columna de tiempo
for i in range(4, m_row + 1):
    cell_obj = sheet_obj.cell(row = i, column = 1)
    tiempo.append(cell_obj.value)

edad=25
BPM=60
PS=120
PD=80
Ruido="si"



#para 60BPM
if BPM==60:

#PS=120 PD=80
    if PS==120 and PD==80:
        for i in range(4, m_row + 1):
            cell_obj = sheet_obj.cell(row = i, column = 2)
            #PM= arr.array('d',[cell_obj.value])
            PM.append(cell_obj.value)

#PS=80 PD=48
    if PS==80 and PD==48:
        for i in range(4, m_row + 1):
            cell_obj = sheet_obj.cell(row = i, column = 4)
            PM.append(cell_obj.value)

#PS=200 PD=140
    if PS==200 and PD==140:
        for i in range(4, m_row + 1):
            cell_obj = sheet_obj.cell(row = i, column = 6)
            PM.append(cell_obj.value)

#para 95BPM
if BPM==95:
#PS=120 PD=80
    if PS==120 and PD==80:
        for i in range(4, m_row + 1):
            cell_obj = sheet_obj.cell(row = i, column = 8)
            PM.append(cell_obj.value)
#PS=80 PD=48
    if PS == 80 and PD == 48:
        for i in range(4, m_row + 1):
            cell_obj = sheet_obj.cell(row = i, column = 10)
            PM.append(cell_obj.value)
#PS=200 PD=140
    if PS==200 and PD==140:
        for i in range(4, m_row + 1):
            cell_obj = sheet_obj.cell(row = i, column = 12)
            PM.append(cell_obj.value)

#para 100BPM
if BPM==100:
#PS=120 PD=80
    if PS==120 and PD==80:
        for i in range(4, m_row + 1):
            cell_obj = sheet_obj.cell(row = i, column = 14)
            PM.append(cell_obj.value)
#PS=80 PD=48
    if PS==80 and PD==48:
        for i in range(4, m_row + 1):
            cell_obj = sheet_obj.cell(row = i, column = 16)
            PM.append(cell_obj.value)
#PS=200 PD=140
    if PS==200 and PD==140:
        for i in range(4, m_row + 1):
            cell_obj = sheet_obj.cell(row = i, column = 18)
            PM.append(cell_obj.value)

#Serial

i=0
#t="a6012080si253623\r\n"


while i<=len(PM)-1:
    ######BPM 60
    #BPM 60 PS==120 and PD==80
    if edad>=10 and PS==120 and PD==80 and BPM==60:
        valor = str(BPM) + str(PS) + str(PD) + Ruido + str(edad) + str(int(PM[i]*100))
        t = "a" + valor + "\r\n"  # 0=t, 1:2=BPM, 3:5=PS, 6:7=PD, 8:9=Ruido, 10:11=edad, 12:...

    if edad<10 and PS==120 and PD==80 and BPM==60:
        valor = str(BPM) + str(PS) + str(PD) + Ruido + str(edad) + str(int(PM[i]*100))
        t = "A" + valor + "\r\n"  # 0=t, 1:2=BPM, 3:5=PS, 6:7=PD, 8:9=Ruido, 10:10=edad, 11:...

    # BPM 60 PS==80 and PD==48
    if edad>=10 and PS==80 and PD==48 and BPM==60:
        valor = str(BPM) + str(PS) + str(PD) + Ruido + str(edad) + str(int(PM[i]*100))
        t = "b" + valor + "\r\n"  # 0=t, 1:2=BPM, 3:4=PS, 5:6=PD, 7:8=Ruido, 9:10=edad, 11:...
    if edad<10 and PS==80 and PD==48 and BPM==60:
        valor = str(BPM) + str(PS) + str(PD) + Ruido + str(edad) + str(int(PM[i]*100))
        t = "B" + valor + "\r\n"   # 0=t, 1:2=BPM, 3:4=PS, 5:6=PD, 7:8=Ruido, 9:9=edad, 10:...

    #BPM 60 PS==200 and PD==140
    if edad>=10 and PS==200 and PD==140 and BPM==60:
        valor = str(BPM) + str(PS) + str(PD) + Ruido + str(edad) + str(int(PM[i]*100))
        t = "c" + valor + "\r\n"  # 0=t, 1:2=BPM, 3:5=PS, 6:8=PD, 9:10=Ruido, 11:12=edad, 13:...
    if edad<10 and PS==200 and PD==140 and BPM==60:
        valor = str(BPM) + str(PS) + str(PD) + Ruido + str(edad) + str(int(PM[i]*100))
        t = "C" + valor + "\r\n"  # 0=t, 1:2=BPM, 3:5=PS, 6:8=PD, 9:10=Ruido, 11:11=edad, 13:...

    ###### BPM 95
    #BPM 95 PS==120 and PD==80
    if edad>=10 and PS==120 and PD==80 and BPM==95:
        valor = str(BPM) + str(PS) + str(PD) + Ruido + str(edad) + str(int(PM[i]*100))
        t = "d" + valor + "\r\n"  # 0=t, 1:2=BPM, 3:5=PS, 6:7=PD, 8:9=Ruido, 10:11=edad, 12:...
    if edad<10 and PS==120 and PD==80 and BPM==95:
        valor = str(BPM) + str(PS) + str(PD) + Ruido + str(edad) + str(int(PM[i]*100))
        t = "D" + valor + "\r\n"  # 0=t, 1:2=BPM, 3:5=PS, 6:7=PD, 8:9=Ruido, 10:10=edad, 11:...

    # BPM 60 PS==80 and PD==48
    if edad>=10 and PS==80 and PD==48 and BPM==95:
        valor = str(BPM) + str(PS) + str(PD) + Ruido + str(edad) + str(int(PM[i]*100))
        t = "e" + valor + "\r\n"  # 0=t, 1:2=BPM, 3:4=PS, 5:6=PD, 7:8=Ruido, 9:10=edad, 11:...
    if edad<10 and PS==80 and PD==48 and BPM==95:
        valor = str(BPM) + str(PS) + str(PD) + Ruido + str(edad) + str(int(PM[i]*100))
        t = "E" + valor + "\r\n"   # 0=t, 1:2=BPM, 3:4=PS, 5:6=PD, 7:8=Ruido, 9:9=edad, 10:...

    #BPM 60 PS==200 and PD==140
    if edad>=10 and PS==200 and PD==140 and BPM==95:
        valor = str(BPM) + str(PS) + str(PD) + Ruido + str(edad) + str(int(PM[i]*100))
        t = "f" + valor + "\r\n"  # 0=t, 1:2=BPM, 3:5=PS, 6:8=PD, 9:10=Ruido, 11:12=edad, 13:...
    if edad<10 and PS==200 and PD==140 and BPM==95:
        valor = str(BPM) + str(PS) + str(PD) + Ruido + str(edad) + str(int(PM[i]*100))
        t = "F" + valor + "\r\n"  # 0=t, 1:2=BPM, 3:5=PS, 6:8=PD, 9:10=Ruido, 11:11=edad, 12:...

    #######BPM 100
    #BPM 100 PS==120 and PD==80
    if edad>=10 and PS==120 and PD==80 and BPM==100:
        valor = str(BPM) + str(PS) + str(PD) + Ruido + str(edad) + str(int(PM[i]*100))
        t = "g" + valor + "\r\n"  # 0=t, 1:3=BPM, 4:6=PS, 7:8=PD, 9:10=Ruido, 11:12=edad, 13:...
    if edad<10 and PS==120 and PD==80 and BPM==100:
        valor = str(BPM) + str(PS) + str(PD) + Ruido + str(edad) + str(int(PM[i]*100))
        t = "G" + valor + "\r\n"  # 0=t, 1:3=BPM, 4:6=PS, 7:8=PD, 9:10=Ruido, 11:11=edad, 12:...

    # BPM 100 PS==80 and PD==48
    if edad>=10 and PS==80 and PD==48 and BPM==100:
        valor = str(BPM) + str(PS) + str(PD) + Ruido + str(edad) + str(int(PM[i]*100))
        t = "h" + valor + "\r\n"  # 0=t, 1:3=BPM, 4:5=PS, 6:7=PD, 8:9=Ruido, 10:11=edad, 12:...
    if edad<10 and PS==80 and PD==48 and BPM==100:
        valor = str(BPM) + str(PS) + str(PD) + Ruido + str(edad) + str(int(PM[i]*100))
        t = "H" + valor + "\r\n"   # 0=t,1:3=BPM, 4:5=PS, 6:7=PD, 8:9=Ruido, 10:10=edad, 11:...

    #BPM 100 PS==200 and PD==140
    if edad>=10 and PS==200 and PD==140 and BPM==100:
        valor = str(BPM) + str(PS) + str(PD) + Ruido + str(edad) + str(int(PM[i]*100))
        t = "i" + valor + "\r\n"  # 0=t, 1:3=BPM, 4:6=PS, 7:9=PD, 10:11=Ruido, 12:13=edad, 14:...
    if edad<10 and PS==200 and PD==140 and BPM==100:
        valor = str(BPM) + str(PS) + str(PD) + Ruido + str(edad) + str(int(PM[i]*100))
        t = "I" + valor + "\r\n"  # 0=t, 1:3=BPM, 4:6=PS, 7:9=PD, 10:11=Ruido, 12:12=edad, 13:...

    arduino.write(t.encode())

    arduino.flushInput()
    print(i)
    i = i + 1
    #print('aa')
    while True:
        try:
            ser_bytes = arduino.readline()
            #print(ser_bytes)
            #print(ser_bytes[0:len(ser_bytes) - 2].decode("utf-8"))
            decoded_bytes = float(ser_bytes[0:len(ser_bytes) - 2].decode("utf-8"))
            #print(decoded_bytes)
            PMr.append(decoded_bytes)

            break
        except:
            print("Keyboard Interrupt")
            break

print(len(PM))
print(len(PMr))
plt.plot(tiempo,PM,'o')
plt.plot(tiempo,PMr,'-')
plt.legend(['PM', 'PM+Ruido'])
plt.xlabel('tiempo(s)')
plt.ylabel('Presión(mmHg)')

plt.show()
